import openpyxl


def readData(rowvalue,columnvalue):
    book = openpyxl.load_workbook("C:\\Users\\Aksha\\OneDrive\\Desktop\\TestData.xlsx")
    sheet = book.active
    if "LoginPage" in book.sheetnames:
        sheet = book.worksheets[0]
        data = sheet.cell(row=rowvalue, column=columnvalue).value
    return data
print(readData(4,1))